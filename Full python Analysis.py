import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

# Load data
file_path = 'D:/jobs/project (1)/dataset.xlsx'  # Replace with the actual path to the data file
campaign_data = pd.read_excel(file_path)

# Metrics Calculation
campaign_data['Conversion_Rate'] = campaign_data['Conversions'] / campaign_data['Clicks']
campaign_data['CPC'] = campaign_data['Total_Spend'] / campaign_data['Clicks']
campaign_data['CPA'] = campaign_data['Total_Spend'] / campaign_data['Conversions']
campaign_data['ROAS'] = campaign_data['Revenue_Generated'] / campaign_data['Total_Spend']

# Create Excel Workbook
wb = Workbook()
ws = wb.active
ws.title = "Campaign Analysis Data"

# Add DataFrame to Excel
for r in dataframe_to_rows(campaign_data, index=False, header=True):
    ws.append(r)

# Create Figures Folder
import os
os.makedirs("figures", exist_ok=True)

# KPI Summary
total_impressions = campaign_data['Impressions'].sum()
total_clicks = campaign_data['Clicks'].sum()
total_conversions = campaign_data['Conversions'].sum()
total_spend = campaign_data['Total_Spend'].sum()
total_revenue = campaign_data['Revenue_Generated'].sum()
average_ctr = total_clicks / total_impressions
average_roas = total_revenue / total_spend

# Save KPIs to Excel
kpi_sheet = wb.create_sheet("KPIs")
kpi_sheet.append(["Metric", "Value"])
kpi_data = [
    ("Total Impressions", total_impressions),
    ("Total Clicks", total_clicks),
    ("Total Conversions", total_conversions),
    ("Total Spend ($)", total_spend),
    ("Total Revenue ($)", total_revenue),
    ("Average CTR (%)", average_ctr * 100),
    ("Average ROAS", average_roas)
]
for kpi, value in kpi_data:
    kpi_sheet.append([kpi, value])

# Campaign Performance Overview
plt.figure(figsize=(14, 8))
sns.barplot(data=campaign_data, x='Campaign_Name', y='ROAS', hue='Marketing_Channel')
plt.xticks(rotation=45, ha="right")
plt.title("Campaign ROAS by Marketing Channel")
plt.savefig("figures/campaign_roas.png")
plt.close()

# Channel Performance Breakdown
channel_data = campaign_data.groupby("Marketing_Channel").sum(numeric_only=True)
fig, ax = plt.subplots(1, 3, figsize=(18, 6))

# Spend by Channel
ax[0].pie(channel_data['Total_Spend'], labels=channel_data.index, autopct='%1.1f%%', startangle=90)
ax[0].set_title("Total Spend by Channel")

# Impressions by Channel
ax[1].pie(channel_data['Impressions'], labels=channel_data.index, autopct='%1.1f%%', startangle=90)
ax[1].set_title("Total Impressions by Channel")

# Conversions by Channel
ax[2].pie(channel_data['Conversions'], labels=channel_data.index, autopct='%1.1f%%', startangle=90)
ax[2].set_title("Total Conversions by Channel")

plt.savefig("figures/channel_performance.png")
plt.close()

# Demographics Insights
fig, axs = plt.subplots(1, 3, figsize=(18, 6))

# Age Group Analysis
age_group_data = campaign_data.groupby("Age_Group").sum(numeric_only=True)
axs[0].bar(age_group_data.index, age_group_data['Conversions'])
axs[0].set_title("Conversions by Age Group")
axs[0].set_xlabel("Age Group")
axs[0].set_ylabel("Conversions")

# Gender Analysis
gender_data = campaign_data.groupby("Gender").sum(numeric_only=True)
axs[1].bar(gender_data.index, gender_data['Conversions'], color=['blue', 'pink'])
axs[1].set_title("Conversions by Gender")
axs[1].set_xlabel("Gender")
axs[1].set_ylabel("Conversions")

# Location Analysis
location_data = campaign_data.groupby("Location").sum(numeric_only=True)
axs[2].bar(location_data.index, location_data['Conversions'], color='purple')
axs[2].set_title("Conversions by Location")
axs[2].set_xlabel("Location")
axs[2].set_ylabel("Conversions")

plt.savefig("figures/demographics_insights.png")
plt.close()

# Time-Based Trend Analysis
campaign_data['Start_Date'] = pd.to_datetime(campaign_data['Start_Date'])
campaign_data.set_index('Start_Date', inplace=True)

# Weekly Trends
weekly_data = campaign_data.resample('W').sum(numeric_only=True)
fig, ax = plt.subplots(figsize=(14, 6))
ax.plot(weekly_data.index, weekly_data['Impressions'], label='Impressions')
ax.plot(weekly_data.index, weekly_data['Clicks'], label='Clicks')
ax.plot(weekly_data.index, weekly_data['Conversions'], label='Conversions')
ax.set_title("Weekly Campaign Metrics Over Time")
ax.set_xlabel("Date")
ax.set_ylabel("Counts")
ax.legend()

plt.savefig("figures/weekly_trends.png")
plt.close()

# Insert images into Excel
image_files = ["campaign_roas.png", "channel_performance.png", "demographics_insights.png", "weekly_trends.png"]
for image_file in image_files:
    img = Image(f"figures/{image_file}")
    ws = wb.create_sheet(image_file.split('.')[0])
    ws.add_image(img, "A1")

# Save final Excel file
final_path = "campaign_analysis_report.xlsx"
wb.save(final_path)
print(f"Excel file saved at {final_path}")
